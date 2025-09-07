from flask import Flask, request, jsonify
from flask_cors import CORS
import mysql.connector
from openpyxl import Workbook, load_workbook
import os

app = Flask(__name__)
CORS(app)

db_config = {
    'host': 'localhost',
    'user': 'root',
    'password': 'rakshit@123',
    'database': 'portfolio'
}

EXCEL_FILE = r"C:\Users\Rakshit\OneDrive\Desktop\websites\portfolio\project\Book1.xlsx"

def save_to_excel(name, email, message):
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(['Name', 'Email', 'Message'])
        wb.save(EXCEL_FILE)
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([name, email, message])
    wb.save(EXCEL_FILE)

@app.route('/contact', methods=['POST'])
def contact():
    data = request.get_json()
    name = data.get('name')
    email = data.get('email')
    message = data.get('message')

    try:
        # Save to MySQL
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor()
        cursor.execute(
            "INSERT INTO info (name, email, message) VALUES (%s, %s, %s)",
            (name, email, message)
        )
        conn.commit()
        cursor.close()
        conn.close()

        # Save to Excel
        save_to_excel(name, email, message)

        return jsonify({'status': 'success', 'message': 'Form received and saved!'})
    except Exception as e:
        print("Error:", e)
        return jsonify({'status': 'error', 'message': str(e)}), 500

if __name__ == '__main__':
    app.run(debug=True)